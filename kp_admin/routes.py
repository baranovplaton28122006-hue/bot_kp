import os
import re
from datetime import datetime, timedelta
from io import StringIO, BytesIO
import csv

from flask import (
    Blueprint, current_app, render_template, request,
    redirect, url_for, send_file, flash, Response, abort
)
from flask_login import login_required, current_user, login_user, logout_user

from . import db
from .models import KPFile, Lead, User
from .utils import parse_kp_file_meta

# NEW: для Excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

bp = Blueprint("kp", __name__)

# ---------- helpers ----------

def uploads_folder() -> str:
    return os.path.abspath(
        os.path.join(current_app.root_path, "..", current_app.config["UPLOAD_FOLDER"])
    )

def paginate(query, per_page: int = 20):
    page = max(int(request.args.get("page", 1) or 1), 1)
    total = query.count()
    pages = (total + per_page - 1) // per_page
    items = query.offset((page - 1) * per_page).limit(per_page).all()
    return items, page, pages, total

# телефон из имени файла: KP_79061419500_20250909_1.html
_FN_PHONE = re.compile(r"^KP_(\d{11,12})_", re.IGNORECASE)

def _sanitize_username(u: str | None) -> str | None:
    if not u:
        return None
    u = u.strip().lstrip("@")
    if u.lower() in {"mail", "email", "почта"}:
        return None
    return u if re.fullmatch(r"[A-Za-z0-9_]{3,32}", u) else None

def _fallback_from_filename(filename: str) -> dict:
    m = _FN_PHONE.match(filename or "")
    return {"phone": ("+" + m.group(1)) if m else None}

def _parse_date_any(s: str | None) -> datetime | None:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    m = re.fullmatch(r"(\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        d, mo = map(int, m.groups())
        today = datetime.utcnow()
        return datetime(today.year, mo, d)
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", s)
    if m:
        d, mo, y = m.groups()
        d, mo, y = int(d), int(mo), int(y)
        if y < 100:
            y += 2000
        return datetime(y, mo, d)
    return None

def get_or_create_lead(meta: dict) -> Lead:
    q = Lead.query
    if meta.get("chat_id"):
        q = q.filter(Lead.tg_user_id == str(meta["chat_id"]))
    elif meta.get("phone"):
        q = q.filter(Lead.phone == meta["phone"])
    else:
        q = q.filter(Lead.username == (meta.get("username") or ""))

    lead = q.first()
    if not lead:
        lead = Lead(
            tg_user_id=str(meta.get("chat_id") or ""),
            username=meta.get("username"),
            phone=meta.get("phone"),
            name=meta.get("name"),
            email=None,
        )
        db.session.add(lead)
        db.session.flush()
    else:
        if meta.get("username") and not lead.username:
            lead.username = meta["username"]
        if meta.get("phone") and not lead.phone:
            lead.phone = meta["phone"]
        if meta.get("name") and not lead.name:
            lead.name = meta["name"]
    return lead

def rescan_new_files(silent: bool = True) -> int:
    folder = uploads_folder()
    os.makedirs(folder, exist_ok=True)

    all_files = [f for f in os.listdir(folder) if f.lower().endswith((".html", ".htm"))]
    added = 0

    for name in sorted(all_files):
        if KPFile.query.filter_by(filename=name).first():
            continue
        path = os.path.join(folder, name)

        try:
            meta = (parse_kp_file_meta(path) or {})
        except Exception:
            meta = {}

        fb = _fallback_from_filename(name)
        if fb.get("phone") and not meta.get("phone"):
            meta["phone"] = fb["phone"]
        meta["username"] = _sanitize_username(meta.get("username"))

        lead = get_or_create_lead({
            "chat_id": meta.get("chat_id"),
            "username": meta.get("username"),
            "phone": meta.get("phone"),
            "name": meta.get("name"),
        })

        db.session.add(KPFile(
            filename=name,
            mimetype="text/html",
            phone=meta.get("phone"),
            chat_id=str(meta.get("chat_id") or ""),
            username=meta.get("username"),
            name=meta.get("name"),
            created_at=datetime.utcnow(),
            lead=lead,
        ))
        added += 1

    if added:
        db.session.commit()

    if not silent:
        flash(f"Сканирую: {uploads_folder()}", "info")
        flash(f"Найдено HTML: {len(all_files)}, добавлено: {added}", "success" if added else "warning")
        if len(all_files) == 0:
            flash("В папке нет *.html. Положи файлы и повтори.", "warning")
    return added

# ---------- routes ----------

@bp.route("/", endpoint="root")
def index_root():
    if current_user.is_authenticated:
        return redirect(url_for("kp.kp_list"))
    return redirect(url_for("kp.login"))

@bp.route("/login", methods=["GET", "POST"], endpoint="login")
def login():
    if request.method == "GET":
        return render_template("login.html")
    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""
    if not email or not password:
        flash("Введите email и пароль", "danger")
        return render_template("login.html"), 400
    user = User.query.filter_by(email=email).first()
    if not user or not user.check_password(password):
        flash("Неверный email или пароль", "danger")
        return render_template("login.html"), 401
    login_user(user, remember=True)
    return redirect(url_for("kp.kp_list"))

@bp.route("/logout", endpoint="logout")
@login_required
def logout():
    logout_user()
    flash("Вы вышли из аккаунта", "success")
    return redirect(url_for("kp.login"))

@bp.route("/kp", endpoint="kp_list")
@login_required
def kp_list():
    q = KPFile.query.order_by(KPFile.created_at.desc())

    phone = (request.args.get("phone") or "").strip()
    if phone:
        q = q.filter(KPFile.phone.contains(phone))

    d_from = _parse_date_any(request.args.get("date_from"))
    d_to   = _parse_date_any(request.args.get("date_to"))
    if d_from:
        q = q.filter(KPFile.created_at >= d_from)
    if d_to:
        q = q.filter(KPFile.created_at < (d_to + timedelta(days=1)))

    items, page, pages, total = paginate(q, per_page=20)

    now = datetime.utcnow()
    today_start = datetime.combine(now.date(), datetime.min.time())
    stat = {
        "today": KPFile.query.filter(KPFile.created_at >= today_start).count(),
        "w7":    KPFile.query.filter(KPFile.created_at >= now - timedelta(days=7)).count(),
        "w30":   KPFile.query.filter(KPFile.created_at >= now - timedelta(days=30)).count(),
        "w90":   KPFile.query.filter(KPFile.created_at >= now - timedelta(days=90)).count(),
        "all":   KPFile.query.count(),
    }

    to_iso = lambda d: d.strftime("%Y-%m-%d") if d else ""
    return render_template(
        "admin/kp_list.html",
        items=items, page=page, pages=pages, total=total, stat=stat,
        date_from_iso=to_iso(d_from), date_to_iso=to_iso(d_to),
        phone_value=phone,
    )

@bp.route("/kp/rescan", methods=["POST", "GET"], endpoint="kp_rescan")
@login_required
def kp_rescan():
    rescan_new_files(silent=False)
    return redirect(url_for("kp.kp_list", **request.args))

# ---- CSV (улучшено: ; и заголовки на русском) ----
@bp.route("/kp/csv", endpoint="kp_csv")
@login_required
def kp_csv():
    q = KPFile.query.order_by(KPFile.created_at.desc())

    phone = (request.args.get("phone") or "").strip()
    if phone:
        q = q.filter(KPFile.phone.contains(phone))

    d_from = _parse_date_any(request.args.get("date_from"))
    d_to   = _parse_date_any(request.args.get("date_to"))
    if d_from:
        q = q.filter(KPFile.created_at >= d_from)
    if d_to:
        q = q.filter(KPFile.created_at < (d_to + timedelta(days=1)))

    si = StringIO()
    w = csv.writer(si, delimiter=';')
    w.writerow(["Username", "Телефон", "Файл", "Дата"])
    for k in q.all():
        w.writerow([
            k.username or "",
            k.phone or "",
            k.filename or "",
            (k.created_at.strftime("%Y-%m-%d %H:%M") if k.created_at else "")
        ])

    return Response(
        si.getvalue().encode("utf-8-sig"),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=kp.csv"}
    )

# ---- XLSX красивый ----
@bp.route("/kp/xlsx", endpoint="kp_xlsx")
@login_required
def kp_xlsx():
    q = KPFile.query.order_by(KPFile.created_at.desc())

    phone = (request.args.get("phone") or "").strip()
    if phone:
        q = q.filter(KPFile.phone.contains(phone))

    d_from = _parse_date_any(request.args.get("date_from"))
    d_to   = _parse_date_any(request.args.get("date_to"))
    if d_from:
        q = q.filter(KPFile.created_at >= d_from)
    if d_to:
        q = q.filter(KPFile.created_at < (d_to + timedelta(days=1)))

    items = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "КП"

    # Шапка
    headers = ["Username", "Телефон", "Файл", "Дата", "Просмотр", "Скачать"]
    ws.append(headers)

    # Стили шапки
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4F81BD")
    thin = Side(border_style="thin", color="D9D9D9")
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center", horizontal="center")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    base_url = request.host_url.rstrip("/")

    # Данные
    for k in items:
        preview_url = url_for("kp.kp_preview", id=k.id, _external=True)
        download_url = url_for("kp.kp_download", id=k.id, _external=True)
        row = [
            k.username or "",
            k.phone or "",
            k.filename or "",
            (k.created_at.strftime("%Y-%m-%d %H:%M") if k.created_at else ""),
            "Просмотр",
            "Скачать",
        ]
        ws.append(row)
        r = ws.max_row
        # Гиперссылки
        ws.cell(row=r, column=5).hyperlink = preview_url
        ws.cell(row=r, column=5).style = "Hyperlink"
        ws.cell(row=r, column=6).hyperlink = download_url
        ws.cell(row=r, column=6).style = "Hyperlink"

    # Ширины колонок
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15

    # Выравнивание/перенос для имени файла
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Автофильтр и фиксация шапки
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    ws.freeze_panes = "A2"

    # Границы для тела
    for r in range(2, ws.max_row+1):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # В буфер
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="kp.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@bp.route("/kp/<int:id>/download", endpoint="kp_download")
@login_required
def kp_download(id: int):
    kpf = KPFile.query.get_or_404(id)
    path = os.path.join(uploads_folder(), kpf.filename)
    if not os.path.exists(path):
        abort(404, "Файл не найден на диске")
    return send_file(path, as_attachment=True, download_name=kpf.filename, mimetype="text/html")

@bp.route("/kp/<int:id>/preview", endpoint="kp_preview")
@login_required
def kp_preview(id: int):
    kpf = KPFile.query.get_or_404(id)
    path = os.path.join(uploads_folder(), kpf.filename)
    if not os.path.exists(path):
        abort(404, "Файл не найден на диске")
    return send_file(path, as_attachment=False, mimetype="text/html")
